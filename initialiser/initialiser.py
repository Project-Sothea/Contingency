import glob
import os

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Side, Border, Alignment

import columns
from columns import DataType
from columns import Types

"""
Initialise an Excel workbook from the patientdata csv file with data validation rules, formatting, and dropdowns.
Takes csv files from patientdata_file_path and types from types_file_path.
Places generated DataSheet.xlsx in the same directory as the script.
"""
class Initialiser:
    def __init__(self, patientdata_file_path, types_filepath):
        """
        Initialize the Validator object with a CSV file path.

        Args:
            csv_folder_path (str): The path to the CSV file to be validated.
        """
        self.csv_folder_path: str = patientdata_file_path
        self.types_filepath: str = types_filepath
        self.df: pd.DataFrame = None
        self.wb = None
        self.ws = None
        self.wb_name = "DataSheet.xlsx"
        self.ws_name = "Sheet1"
        self.types: columns.Types

    def initialise(self):
        """
        Do preliminary initialisation
        """
        try:
            self.types = Types(self.types_filepath)
            print(f"Types '{self.types_filepath}' read successfully!")
        except Exception as e:
            print(f"Error reading types file '{self.types_filepath}': {e}")

    def readCSV(self):
        """
        - Tries to find the latest backup CSV in self.csv_folder_path based on a naming convention.
        - If no matching file is found, defaults to reading any CSV file in the directory.
        - Reads the CSV file into a pandas DataFrame, renames its columns from CSV name -> DataSheet name,
          and initializes the Excel workbook and worksheet.
        """
        try:
            try:
                # Try to find the latest file matching the naming convention
                file_pattern = os.path.join(self.csv_folder_path, "*_patientdata.csv")
                files = glob.glob(file_pattern)

                if not files:
                    raise FileNotFoundError("No files matching the naming convention found.")

                # Extract the timestamp from filenames and sort them
                files_with_timestamps = [
                    (file, filename.split("_patientdata.csv")[0])
                    for file in files
                    for filename in [os.path.basename(file)]
                ]
                latest_file = max(files_with_timestamps, key=lambda x: x[1])[0]
            except Exception:
                # Fallback: Read any CSV file in the directory
                print(
                    "No files matching the naming convention. Falling back to any CSV file in the folder."
                )
                all_files = glob.glob(os.path.join(self.csv_folder_path, "*.csv"))
                if not all_files:
                    raise FileNotFoundError(f"No CSV files found in {self.csv_folder_path}.")
                latest_file = max(all_files, key=os.path.getmtime)  # Use the most recent file

            print(f"CSV file identified: {latest_file}")

            # Read the identified CSV file into a DataFrame
            self.df = pd.read_csv(latest_file)
            print(f"CSV file '{latest_file}' read successfully!")

            # Rename columns of self.df from CSV Name to DataSheet Name
            self.rename_columns()

            # Save DataFrame to an Excel file
            self.df.to_excel(self.wb_name, sheet_name=self.ws_name, index=False)

            # Load and store the workbook
            self.wb = load_workbook(self.wb_name)

            if self.ws_name not in self.wb.sheetnames:
                raise ValueError(f"Sheet '{self.ws_name}' does not exist in the workbook.")

            self.ws = self.wb[self.ws_name]

        except Exception as e:
            print(f"Error reading CSV file: {e}")

    def applyValidationRules(self):
        """
        - Apply data validation rules to an Excel workbook.
        - Disable column renaming #TODO
        - Also adds "Example" Row to workbook #TODO
        """
        try:
            # Add the dv rules listed in DataType enums to the worksheet
            for dt in DataType:
                if dt.validation != None:
                    self.ws.add_data_validation(dt.validation)

            # for each category in types, for each field in the category, apply the validation rule to its datasheet range
            for category in self.types.categories:
                for field in category.fields:
                    if field.datatype.validation != None:
                        field.datatype.validation.add(field.datasheet_range)

            self.wb.save(self.wb_name)
            print("Validation rules applied and saved successfully!")

        except Exception as e:
            print(f"Error applying validation rules: {e}")

    def applyFormatting(self):
        """
        Apply formatting rules to an Excel workbook.
        - Applies a border to first 500 rows
        - Colour first row of cells according to their category
        - Pin first row, and first 3 columns of cells
        - Resizes column width to fit the content.
        - First Row: Triples height, centres text
        - Bold required fields #TODO
        """

        # Apply border to all cells
        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        for row in self.ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=self.ws.max_column):
            for cell in row:
                cell.border = border_style
        print("Borders applied successfully!")

        # Colours first row of cells according to their category
        for category in self.types.categories:
            for field in category.fields:
                custom_fill = PatternFill(start_color=field.colour, end_color=field.colour, fill_type="solid")
                start_cell, end_cell = field.datasheet_range.split(":")
                # Extract the column letters from start_cell and adjust the row to 1
                column_letter = ''.join([char for char in start_cell if char.isalpha()])
                start_cell_adjusted = f"{column_letter}1"

                # Apply the color to the adjusted start cell (now row 1)
                self.ws[start_cell_adjusted].fill = custom_fill
        print("Cell colours applied successfully!")

        # Pin first row, and first 3 columns of cells
        self.ws.freeze_panes = "D2"  # Freeze everything above row 2 and to the left of column D

        # First Row: Triples height, centres text
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(40, max_length + 2)
            self.ws.column_dimensions[column].width = adjusted_width

        self.ws.row_dimensions[1].height = 70
        for cell in self.ws[1]:  # Loop through each cell in the first row
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")


        self.wb.save(self.wb_name)
        print("Formatting applied and saved successfully!")

    def applyConditionalFormatting(self):
        """
        Apply conditional formatting to highlight empty required cells in rows where
        any cell in the same category is filled.
        """
        red_fill = PatternFill(start_color="ea9999", end_color="ea9999",
                               fill_type="solid")  # Red fill for required fields

        for category in self.types.categories:  # Iterate through each category
            # Collect column letters for required fields in this category
            required_columns = [
                ''.join(filter(str.isalpha, f.datasheet_range.split(":")[0]))
                for f in category.fields if f.required  # Only include required fields
            ]
            category_columns = [
                ''.join(filter(str.isalpha, f.datasheet_range.split(":")[0]))
                for f in category.fields
            ]

            for row_number in range(2, 301): # start and end rows
                filled_check = ", ".join([f'ISBLANK(${col}{row_number})' for col in required_columns])
                formula = f'=OR({filled_check})'
                rule = FormulaRule(formula=[formula], fill=red_fill)

                for col in category_columns:
                    self.ws.conditional_formatting.add(f"{col}{row_number}", rule)

        self.wb.save(self.wb_name)
        print("Conditional formatting applied successfully!")

    def rename_columns(self):
        """
        Renames the columns of the existing DataFrame based on a csv_name to datasheet_name mapping.

        Raises:
            ValueError: If self.types or self.df is not initialized.
        """
        # Check if self.types and self.df are initialized
        if self.types is None:
            raise ValueError(
                "The 'types' object is not initialized. Please initialize 'self.types' before renaming columns.")
        if self.df is None:
            raise ValueError(
                "The DataFrame 'self.df' is not initialized. Please load data into 'self.df' before renaming columns.")

        # Generate a mapping from csv_name to datasheet_name
        csv_to_datasheet_map = {}
        for category in self.types.categories:
            for field in category.fields:
                csv_to_datasheet_map[field.csv_name] = field.datasheet_name

        # Rename the columns of the existing DataFrame
        self.df.rename(columns=csv_to_datasheet_map, inplace=True)

if __name__ == "__main__":
    # Path to patient data csv file
    patientdata_file_path = "../cron/backups"  # Replace with your directory path
    types_file_path = "../types.csv"
    datasheet_file_path = "../server/downloads"

    # delete the existing DataSheet.xlsx file
    try:
        os.remove("DataSheet.xlsx")
    except:
        pass

    initialiser = Initialiser(patientdata_file_path, types_file_path)
    initialiser.initialise()
    initialiser.readCSV()
    initialiser.applyValidationRules()
    initialiser.applyFormatting()
    initialiser.applyConditionalFormatting()